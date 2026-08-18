"""Memory-bounded CSV/XLSX parsing with normalized headers."""

from __future__ import annotations

import csv
import io
import math
from collections.abc import Iterable, Iterator, Sequence
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.modules.bulk_imports.errors import (
    BulkImportFileError,
    BulkImportRowLimitError,
    BulkImportTooLargeError,
)

MAX_IMPORT_BYTES = 2 * 1024 * 1024
MAX_IMPORT_ROWS = 500
SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}

ParsedRow = tuple[int, dict[str, object]]


def _normalized_headers(values: Sequence[object]) -> list[str]:
    headers = [str(value).strip().lower() if value is not None else "" for value in values]
    if not headers or any(not header for header in headers):
        raise BulkImportFileError("Every import column must have a non-blank header.")
    if len(set(headers)) != len(headers):
        raise BulkImportFileError("Import column headers must be unique.")
    return headers


def _normalized_scalar(header: str, value: object) -> object:
    """Normalize one cell value to what the Pydantic create schemas expect.

    CSV cells are always ``str`` already (the ``csv`` module never returns
    anything else), so this branching only matters for XLSX, where
    ``openpyxl`` hands back native Python types per cell:

    - ``bool`` is left alone — Pydantic's lax ``bool`` parsing already
      accepts a native bool directly, and stringifying it (``"True"``)
      would turn a real boolean into an identifier-shaped string instead.
    - ``int`` becomes its plain decimal string (``12`` -> ``"12"``) so an
      identifier column (``employee_code``, ``roll_number``, a classroom/
      subject ``code``) typed as a spreadsheet number still validates as
      the ``str`` the schema expects.
    - ``float`` is rejected outright if it is ``NaN``/infinite (neither is
      a legitimate identifier value and both would otherwise stringify to
      something like ``"nan"``/``"inf"``), otherwise rendered as a clean
      decimal string with no spurious trailing zeroes — a whole number
      such as ``12.0`` (Excel's default "General" format for a typed
      number) becomes ``"12"``, not ``"12.0"``.
    - Anything else (e.g. a cell that happens to hold a ``datetime``) is
      passed through unchanged, exactly as before — this fix is
      deliberately narrow and does not attempt to coerce arbitrary
      Python objects.

    None of this touches formula evaluation: the workbook is already
    opened with ``data_only=True`` in ``_xlsx_rows``, so a formula cell
    yields its last-cached value (or ``None``), never a formula string.
    """
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            raise BulkImportFileError(
                f"Column '{header}' contains a NaN or infinite numeric value."
            )
        if value.is_integer():
            return str(int(value))
        text = f"{value:.10f}".rstrip("0").rstrip(".")
        return text
    return value


def _normalized_row(headers: Sequence[str], values: Sequence[object]) -> dict[str, object]:
    row: dict[str, object] = {}
    for header, value in zip(headers, values, strict=False):
        if value is None:
            continue
        if isinstance(value, str):
            value = value.strip()
            if not value:
                continue
            row[header] = value
            continue
        row[header] = _normalized_scalar(header, value)
    return row


def _bounded_rows(rows: Iterable[ParsedRow]) -> list[ParsedRow]:
    parsed: list[ParsedRow] = []
    for row_number, row in rows:
        if not row:
            continue
        if len(parsed) >= MAX_IMPORT_ROWS:
            raise BulkImportRowLimitError(MAX_IMPORT_ROWS)
        parsed.append((row_number, row))
    return parsed


def _csv_rows(content: bytes) -> Iterator[ParsedRow]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise BulkImportFileError("CSV imports must be UTF-8 encoded.") from exc

    try:
        reader = csv.reader(io.StringIO(text, newline=""), strict=True)
        raw_headers = next(reader)
        headers = _normalized_headers(raw_headers)
        for row_number, values in enumerate(reader, start=2):
            if len(values) > len(headers):
                raise BulkImportFileError(f"CSV row {row_number} has more values than the header.")
            yield row_number, _normalized_row(headers, values)
    except StopIteration as exc:
        raise BulkImportFileError("The import file must include a header row.") from exc
    except csv.Error as exc:
        raise BulkImportFileError("The CSV file is malformed.") from exc


def _worksheet_rows(values: Iterable[Sequence[object]]) -> Iterator[ParsedRow]:
    iterator = iter(values)
    try:
        headers = _normalized_headers(next(iterator))
    except StopIteration as exc:
        raise BulkImportFileError("The import file must include a header row.") from exc
    for row_number, row_values in enumerate(iterator, start=2):
        yield row_number, _normalized_row(headers, row_values)


def _xlsx_rows(content: bytes) -> list[ParsedRow]:
    try:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise BulkImportFileError("The XLSX file is malformed.") from exc

    try:
        worksheet = workbook.active
        if worksheet is None:
            raise BulkImportFileError("The XLSX workbook has no active worksheet.")
        rows = _worksheet_rows(worksheet.iter_rows(values_only=True))
        return _bounded_rows(rows)
    finally:
        workbook.close()


def parse_import_file(*, filename: str, content: bytes) -> list[ParsedRow]:
    if len(content) > MAX_IMPORT_BYTES:
        raise BulkImportTooLargeError(MAX_IMPORT_BYTES)
    if not content:
        raise BulkImportFileError("The import file is empty.")

    extension = Path(filename).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise BulkImportFileError("Only .csv and .xlsx import files are supported.")
    if extension == ".csv":
        return _bounded_rows(_csv_rows(content))
    return _xlsx_rows(content)
